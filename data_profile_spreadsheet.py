from attr import NOTHING
import pandas as pd
import urllib
import pyodbc
import shutil
import openpyxl
from sqlalchemy import true
import yaml
import math
from pathlib import Path
from copy import copy
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PortalExporter import PortalConnector

class DataProfileSpreadsheets(object):
    def __init__(self, census_year, census_product, census_table_code, short_title, long_title):
        try:
            self.census_year = census_year
            self.census_product = census_product
            self.census_table_code = census_table_code
            self.short_title = short_title
            self.long_title = long_title
            self.index_sheet_name = 'Geography Index'
            self.geog_groups = {
                'State_County_MSA': ['State', 'County', 'MSA'],
                'City': ['City'],
                'CDP': ['CDP']            
            }
            self.output_dir = Path('./outputs')

        except Exception as e:
            print(e.args[0])
            raise   
        
    
    def build_metadata(self):
        """
        builds metadata dict from year-specific data_profile.yml and a dataset-specific yaml file.
        For the latter, it looks for one specific to the [table]-[year]-[data product].yml,
            if not found then [table]-[year].yml, 
            if that isn't found then [table].yml
        """

        try:
            self.metadata = {}
            dir_name = Path("./Config")
            fpath = dir_name / "data_profile_{}.yml".format(str(self.census_year))
            if fpath.is_file():
                with open(fpath) as file:
                    y = yaml.load(file, Loader=yaml.FullLoader)
            else:
                with open(dir_name / "data_profile.yml") as file:
                    y = yaml.load(file, Loader=yaml.FullLoader)
            self.metadata['metadata'] = y['root']
            c_table = self.census_table_code.lower()
            c_year = str(self.census_year)
            c_product = self.census_product.lower()
            f_name = "_".join([c_table, c_year, c_product]) + ".yml"
            f_path = dir_name / f_name
            if f_path.is_file():
                with open(f_path) as file:
                    y = yaml.load(file, Loader=yaml.FullLoader)
            elif (dir_name / ("_".join([c_table, c_year]) + ".yml")).is_file():
                with open(dir_name / ("_".join([c_table, c_year]) + ".yml")) as file:
                    y = yaml.load(file, Loader=yaml.FullLoader)
            else:
                with open(dir_name / (c_table + ".yml")) as file:
                    y = yaml.load(file, Loader=yaml.FullLoader)
            self.metadata['general'] = y['general']

        except Exception as e:
            print(e.args[0])
            raise   


    def create_data_profile_df(self):
        try:
            conn_string = "DRIVER={ODBC Driver 17 for SQL Server}; SERVER=AWS-PROD-SQL\Sockeye; DATABASE=Elmer; trusted_connection=yes"
            sql_conn = pyodbc.connect(conn_string)
            sql_statement = "select * from data_portal.census_data_profile({}, '{}', '{}')".format(self.census_year, 
                                                                                                self.census_product, 
                                                                                                self.census_table_code)
            self.df = pd.read_sql(sql=sql_statement, con=sql_conn)
            sql_conn.close()
                
        except Exception as e:
            print(e.args[0])
            raise   
        
    
    def geogs_in_geog_type(self, geog_type):
        try:
            df = self.df
            geogs = df[df['geography_type'] == geog_type].geography_name.unique()
            return geogs
            
        except Exception as e:
            print(e.args[0])
            raise   
    
    def filtered_df(self, geog_type, geog):
        """
        Fiters self.df to geog_type and geog,
        and re-creates the index for the new recordset 
        to include sequential integers
        """
        try:
            df = self.df
            df = df[df['geography_type'] == geog_type]
            df = df[df['geography_name'] == geog]
            df = df[['variable_description', 'estimate', 'margin_of_error', 'estimate_percent', 'margin_of_error_percent', 'depth', 'variable_name', 'category']]
            df.columns=['Subject', 'Estimate', 'Margin of Error', 'Percent', 'Percent Margin of Error', 'depth', 'variable_name', 'category']
            df[['Estimate','Margin of Error', 'Percent', 'Percent Margin of Error']] = df[['Estimate','Margin of Error', 'Percent', 'Percent Margin of Error']].fillna('(x)')
            idx_range = range(0, len(df.index))
            ilist = []
            ilist.extend(idx_range)
            df.index = ilist
            return df
        
        except Exception as e:
            print(e.args[0])
            raise   
    
    
    def format_geo_name(self, geog, geog_type):
        try:
            if geog_type == 'CDP':
                geog = geog.replace(' CDP', '')
            if len(geog) > 30:
                geog = geog[0:30]
            return geog
            
        except Exception as e:
            print(e.args[0])
            raise   
    
    
    def prettify_geog_name(self, geog_name, geog_type):
        try: 
            if geog_type in ('MSA', 'State'):
                name = "{} {}".format(geog_name, geog_type)
            else:
                name = geog_name
            return name
       
        except Exception as e:
            print(e.args[0])
            raise   


    def insert_subheaders(self, ws, df):
        """
        insert line breaks whenever the CATEGORY value changes
        """
        try:
            start_row = 2
            # subheader_dict = self.subheader_vars
            # subhead_dict_len = len(subheader_dict)
            subheader_list_len = len(df.category.unique())
            end_row = len(df.index) + start_row + subheader_list_len
            for row in range(start_row, end_row):
                prev_cell_coord = 'H' + str(row-1)
                cell_coord = 'H' + str(row)
                this_cat = ws['I' + str(row)].value
                prev_cat = ws['I' + str(row-1)].value
                if this_cat != prev_cat and prev_cat is not None and this_cat is not None:
                    var_desc = ws['B' + str(row)].value
                    ws.insert_rows(row)
                    subheader_coord = 'A' + str(row)
                    ws[subheader_coord] = this_cat.upper()
                    ws[subheader_coord].font = Font(bold=True)
                
        except Exception as e:
            #print("Error! cell_coord = {}".format(cell_coord))
            print(e.args[0])
            raise   


    def clear_columns(self, ws, df, col_names):
        try:
            start_row = 1
            # subheader_dict = self.subheader_vars
            # subhead_dict_len = len(subheader_dict)
            subhead_list_len = len(df.category.unique())
            end_row = len(df.index) + start_row + subhead_list_len + 2
            for col_name in col_names:
                for row in range(start_row, end_row):
                    cell_coord = col_name + str(row)
                    ws[cell_coord] = None

        except Exception as e:
            #print("Error! cell_coord = {}".format(cell_coord))
            print(e.args[0])
            raise   
    

    def add_notes_sheet(self, wb):

        def bold_cells(data):
            for c in data:
                c = Cell(ws, column="A", row=1, value=c)
                c.font = Font(bold=True)
                yield c

        try: 
            self.build_metadata()
            md = self.metadata
            cell_width = 100
            ws = wb.create_sheet(title='Notes', index=1)
            d_set = md['general']['dataset']
            ws.append(bold_cells([d_set]))
            ws.append(bold_cells(['Each tab contains a distinct geography.  See the "{}" tab for easy navigation.'.format(self.index_sheet_name)]))
            ws.append([])
            syms = md['metadata']['explanation_of_symbols']
            ws.append(bold_cells(['Explanation of Symbols:']))
            for s in syms:
                ws.append(['',s['expl']])
            gen_notes = md['metadata']['notes_applicable_to_all_census_data_profiles']
            ws.append([])
            ws.append(bold_cells(['Notes:']))
            for gn in gen_notes:
                ws.append(['',gn['note']])
            notes = md['general']['notes']
            for n in notes:
                ws.append(['',n['note']])
            col_widths = {'A': 2, 'B':cell_width}
            for cw in col_widths.keys():
                ws.column_dimensions[cw].width = col_widths[cw]
                
            # set row and cell properties
            for row in ws[ws.dimensions]:
                cell = row[1]
                height_val = math.ceil(len(cell.value)/ cell_width) * 15 if cell.value is not None else 15
                ws.row_dimensions[cell.row].height = height_val
                cell.alignment = Alignment(wrap_text=True)
                cell.fill = PatternFill(fill_type='solid', start_color='D7E4BC')
                row[0].fill = PatternFill(fill_type='solid', start_color='D7E4BC')
            

        except Exception as e:
            print(e.args[0])
            raise   

    def get_num_format(self, cell_val):
        try: 
            if (cell_val % 0.1) > 0:
                num_format = '#,##0.00'
            elif (cell_val % 1) > 0:
                num_format = '#,##0.0'
            else:
                num_format = "#,##0"
        except Exception as e:
            print(e.args[0])
            raise   

    def geog_ws(self, geog_type, geog_name, wb):
        """
        Create a new worksheet in the workbook (wb), 
            and fill it with the data profile data for this geog_name.
            Then format the page to make it pretty.
        """
        def is_float(val):
            try:
                num = float(val)
            except ValueError:
                return False
            return True

        try: 
            df = self.df
            data_profile_name = self.long_title
            header_row_size = 1
            f_geog_name = self.format_geo_name(geog_name, geog_type)
            ws = wb.create_sheet(title=f_geog_name)
            filtereddf = self.filtered_df(geog_type, geog_name)
            #finaldf = filtereddf.drop(['depth', 'variable_name'], axis=1)
            for r in dataframe_to_rows(filtereddf, index=False, header=True):
                #self.add_section_header(r, filtereddf, ws)
                r.insert(0,'')
                ws.append(r)

            #format worksheet body
            col_widths = {'A':2, 'B':75, 'C':15, 'D':15, 'E':15, 'F':15}
            for cw in col_widths.keys():
                ws.column_dimensions[cw].width = col_widths[cw]
            max_row = len(filtereddf.index) + header_row_size + 1
            #col_styles = {'C':{'align':'center', 'num':'0,##'}, 
            col_styles = {'C':{'align':'center', 'num':'General'}, 
                        'D':{'align':'center', 'num':'General'},
                        'E':{'align':'center', 'num':'#0.0'},
                        'F':{'align':'center', 'num':'#0.0'}}
            # dec_vars = self.decimal_estimate_vars[self.census_table_code]
            for r in range(header_row_size + 1, max_row):
                if self.census_table_code == 'DP05':
                    indent_level = (filtereddf.depth[r - 2]) * 2
                else:
                    indent_level = (filtereddf.depth[r - 2] - 1) * 2
                ws['B'+str(r)].alignment = Alignment(indent = indent_level)
                for c in col_styles.keys():
                    cell = ws[c+str(r)]
                    cell.alignment = Alignment(horizontal=col_styles[c]['align']) 
                    #if filtereddf.variable_name[r-2] in dec_vars:
                    if is_float(cell.value):
                        if cell.value > 100:
                            cell.number_format = '#,##0'
                    else:
                        cell.number_format = col_styles[c]['num']

            #format worksheet header
            ws.row_dimensions[1].height = 30
            cols = ['A','B','C','D','E','F']
            for c in cols:
                cell = ws[c+str(1)]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal='center')
                cell.fill = PatternFill(fill_type='solid', start_color='D7E4BC')
                header_border = Side(border_style="thin", color="000000")
                cell.border = Border(top=header_border, 
                                    left=header_border, 
                                    right=header_border, 
                                    bottom=header_border)
            
            
            self.insert_subheaders(ws, filtereddf)
            self.clear_columns(ws, filtereddf, ['G','H','I'])
                
            #insert title row
            ws.insert_rows(1)
            title_cell = ws['A1']
            pretty_geog_name = self.prettify_geog_name(geog_name, geog_type)
            title_cell.value = "{}: {}".format(data_profile_name, pretty_geog_name)
            title_cell.font = Font(size=18, bold=True)

            self.merge_cells_a_and_b(ws)
            if self.census_product == 'Decennial' and self.census_year >= 2020:
                self.add_special_subheadings(ws)
            
        except Exception as e:
            print(e.args[0])
            print("r value: {}".format(r))
            raise   
    
    

    def add_special_subheadings(self, ws):
        try:
            # print("I'm in add_special_subheadings.  max_row = {}".format(str(ws.max_row)))
            for row in range(4, ws.max_row + 1):
                cur_cell = ws.cell(row=row, column=2)
                above_cell = ws.cell(row=row - 1, column=2)
                test_str = '16 years and over'
                head_str = '  Selected Age Categories'
                if cur_cell.value and above_cell.value:
                    if (cur_cell.value == test_str) and (above_cell.value != head_str):
                        # print("current cell: {}".format(cur_cell.value))
                        ws.insert_rows(row)
                        new_cell = ws.cell(row=row, column=2)
                        new_cell.value = head_str

        except Exception as e:
            print("exception in add_special_subheadings {}", format(e.args[0]))
            raise   
    
    
    def merge_cells_a_and_b(self, ws):
        try:
            cell_a = ws['A2']
            cell_b = ws['B2']
            cell_a.value =  cell_b.value
            cell_a.border =  copy(cell_b.border)
            cell_a.alignment =  copy(cell_b.alignment)
            cell_a.fill =  copy(cell_b.fill)
            ws.merge_cells("A2:B2")

        except Exception as e:
            print(e.args[0])
            raise   
    
    
    def add_links_to_index(self, wb, geogs, geog_types, file_name):
        """
        Add a new worksheet, with hyperlinks to all the other sheets.
        """
        try:
            index_sheet_name = self.index_sheet_name
            index_sheet = wb[index_sheet_name]
            index_sheet.column_dimensions['A'].width = 100
            title_cell = index_sheet['A1']
            title_cell.value = self.long_title
            title_cell.font = Font(size=18, bold=True)
            header_cell = index_sheet['A2']
            header_cell.value = "For summary data for a specific geography, click a link below:"
            index_row = 3
            for geog_type in geog_types:
                geogs = self.geogs_in_geog_type(geog_type)
                for g in geogs:
                    target_sheet_name = self.format_geo_name(g, geog_type)
                    pretty_geog_name = self.prettify_geog_name(g, geog_type)
                    link_ws = wb[index_sheet_name]
                    link = file_name + "#'{}'!A1".format(target_sheet_name)
                    cell = link_ws.cell(row=index_row, column=1)
                    cell.hyperlink = link
                    cell.value = pretty_geog_name
                    cell.style = "Hyperlink"
                    index_row += 1 

        except Exception as e:
            print(e.args[0])
            raise   
    
    
    def create_wb(self, geog_group_name, geog_types):
        try:
            data_profile_name = self.long_title
            filename_stub = self.short_title.replace(" ", "_")
            vintage = self.census_year
            census_table = self.census_table_code
            c_product = self.census_product
            wb = Workbook()
            fname = "{}_{}_{}_{}.xlsx".format(filename_stub, geog_group_name, vintage, c_product)
            fpath = self.output_dir / fname
            index_sheet = wb["Sheet"]
            index_sheet.title = self.index_sheet_name
            for geog_type in geog_types:
                geogs = self.geogs_in_geog_type(geog_type)
                for geo in geogs:
                    self.geog_ws(geog_type, geo, wb)
            self.add_links_to_index(wb, geogs, geog_types, fname)
            self.add_notes_sheet(wb)

            wb.save(filename=fpath)   
            return(fpath)
        
        except Exception as e:
            print(e.args[0])
            raise   

    
    def create_workbooks(self):
        """
        Create Excel workbooks for all geography groups (State-County-MSA, City, and CDP).
        Exports one workbook per geography group.
        """
        try:
            self.create_data_profile_df()
            df = self.df
            # short_title = self.short_title
            vintage = self.census_year
            # census_table = self.census_table_code
            geog_types = df.geography_type.unique()
            geog_groups = {'State': ['County']} ## Testing only!!!  Comment this out for production.
            geog_groups = self.geog_groups
            if self.census_product == 'acs1': # Census does not publish data profiles at CDP level for 1-year data
                geog_groups.pop('CDP')
            for g_group_name in geog_groups.keys():
                data_profile_name = '_by_'.join([self.long_title, g_group_name])
                geog_types = geog_groups[g_group_name]
                filename = self.create_wb(g_group_name, geog_types)
                #self.export_workbook(filename=filename, portal_layer_name=data_profile_name)

        except Exception as e:
            print(e.args[0])
            raise   

    def export_workbook(self, filename, portal_layer_name):
        """
        Send a spreadsheet to Portal.  
        Shares the layer with everyone.
        
        params:
            filename: the local path to the spreadsheet
            portal_layer_name: the name to be published under on Portal
        """
        try:
            config_path = Path('Config') / 'auth.yml'
            with open(config_path) as file:
                auth = yaml.load(file, Loader=yaml.FullLoader)
            portal_conn = PortalConnector(
                portal_username = auth['arc_gis_online']['username'],
                portal_pw=auth['arc_gis_online']['pw']
            )
            gis = portal_conn.gis
            resource_properties = {
                'title': portal_layer_name, 
                'tags': ['data profile']}
            exported = gis.content.add(resource_properties, data=str(filename))
            portal_group = gis.groups.search(query='PSRC Data Portal Content')[0].id
            census_group = gis.groups.search(query='hub- Census & Demographics')[0].id
            exported.share(everyone=True, groups=[portal_group, census_group])

        except Exception as e:
            print(e.args[0])
            raise   